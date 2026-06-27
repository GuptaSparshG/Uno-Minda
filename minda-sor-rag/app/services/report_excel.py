"""Excel report generator (openpyxl).

Workbook layout:
  Sheet 1: Summary
  Sheet 2: All Statements
  Sheet 3+: one per section (name truncated to 31 chars — Excel limit)
"""

from __future__ import annotations

import re

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app.schemas import AnalysisResult, SectionResult

HEADER_FILL = PatternFill("solid", fgColor="1A365D")
HEADER_FONT = Font(bold=True, color="FFFFFF")

CLASS_FILLS = {
    "REQUIREMENT": PatternFill("solid", fgColor="BEE3F8"),
    "RECOMMENDATION": PatternFill("solid", fgColor="FEFCBF"),
    "ASK": PatternFill("solid", fgColor="E9D8FD"),
    "INFORMATIONAL": PatternFill("solid", fgColor="E2E8F0"),
}

SCORE_GOOD = PatternFill("solid", fgColor="C6F6D5")
SCORE_WARN = PatternFill("solid", fgColor="FEFCBF")
SCORE_BAD = PatternFill("solid", fgColor="FED7D7")

STATEMENT_HEADER = [
    "ID",
    "Section",
    "Text",
    "Classification",
    "ISO Category",
    "Obligation",
    "Score",
    "Violated Rules",
    "Suggested Action",
]

COL_WIDTHS = [10, 22, 70, 18, 18, 12, 8, 20, 40]


def generate(result: AnalysisResult, output_path: str) -> None:
    wb = Workbook()

    _build_summary_sheet(wb, result)
    _build_all_statements_sheet(wb, result)
    for section in result.sections:
        _build_section_sheet(wb, section)

    wb.save(output_path)


def _build_summary_sheet(wb: Workbook, result: AnalysisResult) -> None:
    ws = wb.active
    ws.title = "Summary"

    ws["A1"] = "SOR Requirements Analysis Report"
    ws["A1"].font = Font(bold=True, size=14)
    ws.merge_cells("A1:E1")

    ws["A3"] = "File:"
    ws["B3"] = result.filename
    ws["A4"] = "Total Statements:"
    ws["B4"] = result.total_statements
    ws["A5"] = "Total Sections:"
    ws["B5"] = result.total_sections

    ws["A7"] = "Classification Breakdown"
    ws["A7"].font = Font(bold=True)

    totals = {"REQUIREMENT": 0, "RECOMMENDATION": 0, "ASK": 0, "INFORMATIONAL": 0}
    for sec in result.sections:
        totals["REQUIREMENT"] += sec.requirements
        totals["RECOMMENDATION"] += sec.recommendations
        totals["ASK"] += sec.asks
        totals["INFORMATIONAL"] += sec.informational

    rows = [
        ("Requirements", totals["REQUIREMENT"]),
        ("Recommendations", totals["RECOMMENDATION"]),
        ("Asks", totals["ASK"]),
        ("Informational", totals["INFORMATIONAL"]),
    ]
    for i, (label, value) in enumerate(rows, start=8):
        ws[f"A{i}"] = label
        ws[f"B{i}"] = value

    ws["A13"] = "Section Summary"
    ws["A13"].font = Font(bold=True)

    headers = ["Section", "Total", "Requirements", "Asks", "Avg Score"]
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=14, column=col, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT

    for i, sec in enumerate(result.sections, start=15):
        ws.cell(row=i, column=1, value=sec.section_name)
        ws.cell(row=i, column=2, value=sec.total)
        ws.cell(row=i, column=3, value=sec.requirements)
        ws.cell(row=i, column=4, value=sec.asks)
        ws.cell(row=i, column=5, value=sec.avg_quality_score)

    for col, width in enumerate([28, 10, 14, 10, 12], start=1):
        ws.column_dimensions[get_column_letter(col)].width = width


def _build_all_statements_sheet(wb: Workbook, result: AnalysisResult) -> None:
    ws = wb.create_sheet("All Statements")
    _write_statement_header(ws)
    row = 2
    for sec in result.sections:
        for stmt in sec.statements:
            _write_statement_row(ws, row, stmt)
            row += 1
    _finalize_statement_sheet(ws, row - 1)


def _build_section_sheet(wb: Workbook, section: SectionResult) -> None:
    title = _safe_sheet_title(section.section_name)
    # avoid duplicate sheet names
    base = title
    i = 2
    while title in wb.sheetnames:
        suffix = f" ({i})"
        title = (base[: 31 - len(suffix)] + suffix)
        i += 1
    ws = wb.create_sheet(title)

    ws["A1"] = section.section_name
    ws["A1"].font = Font(bold=True, size=12)
    ws.merge_cells("A1:I1")
    ws["A2"] = (
        f"Total: {section.total} | Req: {section.requirements} | "
        f"Rec: {section.recommendations} | Ask: {section.asks} | "
        f"Info: {section.informational} | Avg Score: {section.avg_quality_score}"
    )
    ws.merge_cells("A2:I2")

    header_row = 4
    for col, h in enumerate(STATEMENT_HEADER, start=1):
        cell = ws.cell(row=header_row, column=col, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT

    row = header_row + 1
    for stmt in section.statements:
        _write_statement_row(ws, row, stmt)
        row += 1

    _finalize_statement_sheet(ws, row - 1, header_row=header_row)


def _write_statement_header(ws) -> None:
    for col, h in enumerate(STATEMENT_HEADER, start=1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT


def _write_statement_row(ws, row: int, stmt) -> None:
    ws.cell(row=row, column=1, value=stmt.id)
    ws.cell(row=row, column=2, value=stmt.section)
    text_cell = ws.cell(row=row, column=3, value=stmt.text)
    text_cell.alignment = Alignment(wrap_text=True, vertical="top")

    class_cell = ws.cell(row=row, column=4, value=stmt.classification)
    fill = CLASS_FILLS.get(stmt.classification)
    if fill is not None:
        class_cell.fill = fill

    ws.cell(row=row, column=5, value=stmt.iso_category)
    ws.cell(row=row, column=6, value=stmt.obligation_keyword)

    score_cell = ws.cell(row=row, column=7, value=stmt.quality_score)
    if stmt.quality_score >= 80:
        score_cell.fill = SCORE_GOOD
    elif stmt.quality_score >= 50:
        score_cell.fill = SCORE_WARN
    else:
        score_cell.fill = SCORE_BAD

    ws.cell(row=row, column=8, value=", ".join(stmt.violated_rules))
    ws.cell(row=row, column=9, value=stmt.suggested_action)


def _finalize_statement_sheet(ws, last_row: int, header_row: int = 1) -> None:
    for col, width in enumerate(COL_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)
    if last_row >= header_row:
        end_col = get_column_letter(len(STATEMENT_HEADER))
        ws.auto_filter.ref = f"A{header_row}:{end_col}{max(last_row, header_row)}"


def _safe_sheet_title(name: str) -> str:
    cleaned = re.sub(r"[\\/*?:\[\]]", " ", name).strip() or "Section"
    return cleaned[:31]
